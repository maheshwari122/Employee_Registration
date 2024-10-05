from django.shortcuts import render
from django.core import serializers

# Create your views here.
from django.shortcuts import render
from registration.models import EmployeeRegistrationForm
from .models import EmployeeRegistrationForm
from .serializers import EmployeeRegistrationFormSerializers
from rest_framework.response import Response
from utils import custom_viewsets

from .forms import NameForm


from rest_framework.decorators import *
from rest_framework.viewsets import ModelViewSet
from rest_framework.permissions import IsAdminUser
import openpyxl
from django.http import HttpResponse

        

class EmployeeRegistrationFormViewSet(custom_viewsets.ModelViewSet):
   # permission_classes = [IsAdminUser]
    model = EmployeeRegistrationForm
    queryset = EmployeeRegistrationForm.objects.all()
    serializer_class = EmployeeRegistrationFormSerializers
    list_success_message = "Employee list returend success"
    retrieve_sucess_message = "Employee retrieve returend success" 
    create_success_message = "Employee create the data"
    status_response = 200
    status_code = 200
    response = {
        "status": True,
        "msg": None,
        "data": {}
    }

    @action(detail=False, methods=['POST'])
    def registration_form(request):
        self.response.update({
                "status": 200,
                "msg": 'employee data view featch',
                "data": {}
            })
        return Response(self.response)

@api_view(['GET', 'POST'])
def Registration(request):
    if request.method == 'POST':
        fname = request.POST.get('fname')
        phoneNo = request.POST.get('phoneNo')
        email = request.POST.get('email')
        age = request.POST.get('age')
        gender = request.POST.get('gender')
        date_of_birth = request.POST.get('date_of_birth')
        marital_status = request.POST.get('marital_status')
        education = request.POST.get('education')
        residence_address = request.POST.get('residence_address')
        present_address = request.POST.get('present_address')
        state = request.POST.get('state')
        city = request.POST.get('city')
        country = request.POST.get('country')

        employeeModel = EmployeeRegistrationForm()
        employeeModel.full_name = fname
        employeeModel.phone_no = phoneNo
        employeeModel.email = email
        employeeModel.age = age
        employeeModel.gender = gender
        employeeModel.Date_of_birth = date_of_birth
        employeeModel.marital_status = marital_status
        employeeModel.education = education
        employeeModel.residence_address = residence_address
        employeeModel.present_address = present_address
        employeeModel.state = state
        employeeModel.city = city
        employeeModel.country = country
        
        employeeModel.save()

        allEmployee = EmployeeRegistrationForm.objects.all()
        data = {"employees":allEmployee}
        return render(request, "template/users.html", data)
        
        # print("request happend " , request.POST.get('fname'))

    form = NameForm()
    # return render(request, "template/home.html", {"form": form})    
    return render(request, "template/home.html")        

@api_view(['GET'])
def ExportEmployeeExcel(request, id):
    print("Employee data : ", id)
    employee =serializers.serialize("json", EmployeeRegistrationForm.objects.filter(id=1))
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    title = "Employee_Data_"
    

    header = ['ID', 'Full Name', 'Email', 'Phone', 'Age', 'Gender', 'Marital Status', 'Education', 'Residance Address', 'Present Address', 'State', 'City', 'Country']
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    queryset = EmployeeRegistrationForm.objects.filter(id=id).values_list('id', 'full_name', 'email', 'phone_no', 'age', 'gender', 'marital_status', 'education', 'residence_address', 'present_address', 'state', 'city', 'country')

    for row_num, row in enumerate(queryset, 1):
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num+1, column=col_num)
            cell.value = cell_value
            if col_num == 2:
                title += cell_value
            
    worksheet.title = title
    fileAttachement = 'attachment; filename={fileName}.xlsx'
    attachment  = fileAttachement.format(fileName=title)
    response['Content-Disposition'] = attachment
    workbook.save(response)
    return response